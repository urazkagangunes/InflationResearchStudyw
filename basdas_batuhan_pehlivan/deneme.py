import os
import re
from datetime import datetime
from typing import List, Tuple, Optional

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


BASE_URL = "https://basdasonline.com"
LIST_URL = f"{BASE_URL}/tab-lists.asp"

EXCEL_PATH = "basdas_fiyat_takip.xlsx"
SHEET_NAME = "data"
HEADERS = ["tarih", "grup_id", "urun_adi", "fiyat"]




def ensure_workbook(path: str = EXCEL_PATH):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        if ws.max_row == 0:
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:D1"

    widths = [20, 10, 60, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb, ws


def append_rows(rows):
    if not rows:
        return

    wb, ws = ensure_workbook()


    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):

        existing.add((row[0], row[1], row[2]))

    new_count = 0

    for r in rows:
        key = (r[0], r[1], r[2])

        if key not in existing:
            ws.append(r)
            new_count += 1

    wb.save(EXCEL_PATH)

    print(f"Yeni eklenen satır: {new_count}")


_price_clean_re = re.compile(r"[^\d,\.]")

def parse_price(text: str) -> Optional[float]:
    if not text:
        return None
    t = _price_clean_re.sub("", text.strip())

    if "," in t and "." in t:
        t = t.replace(".", "").replace(",", ".")
    elif "," in t:
        t = t.replace(",", ".")

    try:
        return float(t)
    except:
        return None


def parse_products(html: str):
    soup = BeautifulSoup(html, "html.parser")
    products = []


    cards = soup.select(".urun-kutusu")

    for card in cards:
        name_el = card.select_one("h2 a.kutu-link")
        price_el = card.select_one("div.urun-fiyat span")

        if not name_el or not price_el:
            continue

        name = name_el.get_text(strip=True)
        price = parse_price(price_el.get_text(strip=True))

        if name and price is not None:
            products.append((name, price))

    return products




def main():
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0",
        "Accept": "*/*",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": BASE_URL + "/",
        "Origin": BASE_URL,
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    })


    s.get(BASE_URL + "/", timeout=20)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_rows = []

    for gid in range(0, 101):
        try:
            r = s.post(LIST_URL, data={"grupID": str(gid)}, timeout=25)
            r.encoding = "utf-8"

            products = parse_products(r.text)

            if products:
                print(f"[OK] grupID={gid} → {len(products)} ürün")

                for name, price in products:
                    all_rows.append((now, gid, name, price))

        except Exception as e:
            print(f"[ERR] grupID={gid} → {e}")

    append_rows(all_rows)

    print(f"\nExcel'e yazıldı: {EXCEL_PATH}")
    print(f"Toplam eklenen satır: {len(all_rows)}")


if __name__ == "__main__":
    main()